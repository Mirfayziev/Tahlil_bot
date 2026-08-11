import secrets
from datetime import datetime

from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app
from flask_login import login_required, current_user

from app.extensions import db
from app.decorators import roles_required
from app.security import validate_password_strength
from app.models import (
    ServiceCategory, User, Department, RoleEnum, AuditLog, Building, Employee, Customer,
    ServiceRequest
)

admin_bp = Blueprint("admin", __name__)


def log_action(action, entity, entity_id=None, details=None):
    db.session.add(AuditLog(
        user_id=current_user.id if current_user.is_authenticated else None,
        action=action, entity=entity, entity_id=entity_id, details=details
    ))


# ---------------------------------------------------------------------------
# Xizmat kategoriyalari
# ---------------------------------------------------------------------------
@admin_bp.route("/categories", methods=["GET", "POST"])
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def categories():
    if request.method == "POST":
        cat = ServiceCategory(
            name_uz=request.form["name_uz"],
            name_ru=request.form.get("name_ru"),
            name_en=request.form.get("name_en"),
            description=request.form.get("description"),
            parent_id=request.form.get("parent_id") or None,
            department_id=request.form.get("department_id") or None,
            default_priority=request.form.get("default_priority", "orta"),
            default_sla_hours=int(request.form.get("default_sla_hours", 24)),
        )
        db.session.add(cat)
        log_action("create", "ServiceCategory", details=cat.name_uz)
        db.session.commit()
        flash("Kategoriya qo'shildi.", "success")
        return redirect(url_for("admin.categories"))

    all_categories = ServiceCategory.query.order_by(ServiceCategory.sort_order).all()
    top_level = [c for c in all_categories if c.parent_id is None]
    departments = Department.query.all()
    return render_template("admin/categories.html", categories=all_categories, top_level=top_level,
                            departments=departments)


@admin_bp.route("/categories/<int:cat_id>/toggle", methods=["POST"])
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def toggle_category(cat_id):
    cat = ServiceCategory.query.get_or_404(cat_id)
    cat.is_active = not cat.is_active
    db.session.commit()
    return redirect(url_for("admin.categories"))


@admin_bp.route("/categories/<int:cat_id>/set-department", methods=["POST"])
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def set_category_department(cat_id):
    cat = ServiceCategory.query.get_or_404(cat_id)
    cat.department_id = request.form.get("department_id") or None
    log_action("update", "ServiceCategory", entity_id=cat.id,
               details=f"department_id={cat.department_id}")
    db.session.commit()
    flash(f"{cat.name_uz} — bo'lim yangilandi.", "success")
    return redirect(url_for("admin.categories"))


# ---------------------------------------------------------------------------
# Xodimlar / Ijrochilar
# ---------------------------------------------------------------------------
@admin_bp.route("/staff", methods=["GET", "POST"])
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def staff():
    if request.method == "POST":
        password = request.form.get("password") or secrets.token_urlsafe(9)
        errors = validate_password_strength(password, current_app.config.get("PASSWORD_MIN_LENGTH", 8))
        if errors:
            for e in errors:
                flash(e, "danger")
            return redirect(url_for("admin.staff"))

        user = User(
            full_name=request.form["full_name"],
            username=request.form["username"],
            role=request.form["role"],
            phone=request.form.get("phone"),
            telegram_id=request.form.get("telegram_id") or None,
            department_id=request.form.get("department_id") or None,
            position=request.form.get("position"),
        )
        user.set_password(password)
        db.session.add(user)
        log_action("create", "User", details=user.username)
        db.session.commit()
        flash(f"Xodim yaratildi. Vaqtinchalik parol: {password}", "success")
        return redirect(url_for("admin.staff"))

    users = User.query.order_by(User.created_at.desc()).all()
    departments = Department.query.all()
    buildings = Building.query.all()
    roles = list(RoleEnum)
    return render_template("admin/staff.html", users=users, departments=departments,
                            buildings=buildings, roles=roles)


@admin_bp.route("/staff/<int:user_id>/toggle", methods=["POST"])
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def toggle_staff(user_id):
    user = User.query.get_or_404(user_id)
    user.is_active_flag = not user.is_active_flag
    db.session.commit()
    return redirect(url_for("admin.staff"))


@admin_bp.route("/staff/<int:user_id>/set-departments", methods=["POST"])
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def set_staff_departments(user_id):
    """Ijrochini bir nechta yo'nalishga (bo'limga) biriktirish — masalan
    'Santexnika va Elektr' + 'Konditsioner'."""
    user = User.query.get_or_404(user_id)
    dept_ids = [int(d) for d in request.form.getlist("department_ids")]
    user.departments = Department.query.filter(Department.id.in_(dept_ids)).all() if dept_ids else []
    log_action("update", "User", entity_id=user.id, details=f"department_ids={dept_ids}")
    db.session.commit()
    flash(f"{user.full_name} — yo'nalishlari yangilandi.", "success")
    return redirect(url_for("admin.staff"))


@admin_bp.route("/staff/<int:user_id>/set-department", methods=["POST"])
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def set_staff_department(user_id):
    user = User.query.get_or_404(user_id)
    user.department_id = request.form.get("department_id") or None
    log_action("update", "User", entity_id=user.id, details=f"department_id={user.department_id}")
    db.session.commit()
    flash(f"{user.full_name} — bo'lim (yo'nalish) yangilandi.", "success")
    return redirect(url_for("admin.staff"))


@admin_bp.route("/staff/<int:user_id>/set-buildings", methods=["POST"])
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def set_staff_buildings(user_id):
    """Ijrochini bir nechta binoga biriktirish — masalan Markaziy Apparat + Minor."""
    user = User.query.get_or_404(user_id)
    building_ids = [int(b) for b in request.form.getlist("building_ids")]
    user.buildings = Building.query.filter(Building.id.in_(building_ids)).all() if building_ids else []
    log_action("update", "User", entity_id=user.id, details=f"building_ids={building_ids}")
    db.session.commit()
    flash(f"{user.full_name} — binolari yangilandi.", "success")
    return redirect(url_for("admin.staff"))


# ---------------------------------------------------------------------------
# Bo'limlar
# ---------------------------------------------------------------------------
@admin_bp.route("/departments", methods=["GET", "POST"])
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def departments():
    if request.method == "POST":
        name = request.form["name"].strip()
        existing = Department.query.filter(db.func.lower(Department.name) == name.lower()).first()
        if existing:
            flash(
                f"\"{existing.name}\" nomli bo'lim allaqachon mavjud (ID: {existing.id}). "
                f"Bir xil nomli ikkita bo'lim yaratish kategoriya/ijrochi biriktirishda "
                f"chalkashlikka olib keladi — kerak bo'lsa mavjudini tahrirlang.",
                "danger",
            )
            return redirect(url_for("admin.departments"))

        dep = Department(name=name, description=request.form.get("description"))
        db.session.add(dep)
        db.session.commit()
        flash("Bo'lim qo'shildi.", "success")
        return redirect(url_for("admin.departments"))

    deps = Department.query.all()
    return render_template("admin/departments.html", departments=deps)


@admin_bp.route("/departments/<int:dep_id>/delete", methods=["POST"])
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def delete_department(dep_id):
    dep = Department.query.get_or_404(dep_id)
    # Bog'liq kategoriya/xodimlarni "bog'lanmagan" holatga o'tkazamiz (o'chirish
    # ularni buzmaydi) - aks holda ma'lumotlar bazasi cheklovi buzilar edi.
    ServiceCategory.query.filter_by(department_id=dep.id).update({"department_id": None})
    User.query.filter_by(department_id=dep.id).update({"department_id": None})
    dep.executors = []
    log_action("delete", "Department", entity_id=dep.id, details=dep.name)
    db.session.delete(dep)
    db.session.commit()
    flash(f"\"{dep.name}\" bo'limi o'chirildi.", "success")
    return redirect(url_for("admin.departments"))


@admin_bp.route("/buildings", methods=["GET", "POST"])
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def buildings():
    if request.method == "POST":
        name = request.form["name"].strip()
        existing = Building.query.filter(db.func.lower(Building.name) == name.lower()).first()
        if existing:
            flash(
                f"\"{existing.name}\" nomli bino allaqachon mavjud (ID: {existing.id}). "
                f"Bir xil nomli ikkita bino ijrochi biriktirishda chalkashlikka olib keladi.",
                "danger",
            )
            return redirect(url_for("admin.buildings"))

        max_order = db.session.query(db.func.max(Building.sort_order)).scalar() or 0
        b = Building(name=name, description=request.form.get("description"), sort_order=max_order + 1)
        db.session.add(b)
        db.session.commit()
        flash("Bino qo'shildi.", "success")
        return redirect(url_for("admin.buildings"))

    blds = Building.query.order_by(Building.sort_order, Building.name).all()
    return render_template("admin/buildings.html", buildings=blds)


@admin_bp.route("/buildings/<int:building_id>/reorder", methods=["POST"])
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def reorder_building(building_id):
    b = Building.query.get_or_404(building_id)
    try:
        b.sort_order = int(request.form.get("sort_order", 0))
    except ValueError:
        flash("Tartib raqami butun son bo'lishi kerak.", "danger")
        return redirect(url_for("admin.buildings"))
    db.session.commit()
    flash(f"\"{b.name}\" tartibi yangilandi.", "success")
    return redirect(url_for("admin.buildings"))


@admin_bp.route("/buildings/<int:building_id>/delete", methods=["POST"])
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def delete_building(building_id):
    b = Building.query.get_or_404(building_id)
    ServiceRequest.query.filter_by(building_id=b.id).update({"building_id": None})
    b.executors = []
    log_action("delete", "Building", entity_id=b.id, details=b.name)
    db.session.delete(b)
    db.session.commit()
    flash(f"\"{b.name}\" binosi o'chirildi.", "success")
    return redirect(url_for("admin.buildings"))


def _normalize_pinfl(raw: str) -> str:
    return "".join(ch for ch in (raw or "") if ch.isdigit())


# ---------------------------------------------------------------------------
# Tashkilot xodimlari (PINFL ro'yxati — murojaatchi bot orqali xizmatdan
# foydalanish uchun tekshiriladigan ro'yxat, tizimga kiruvchi Users'dan farqli)
# ---------------------------------------------------------------------------
@admin_bp.route("/employees", methods=["GET", "POST"])
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def employees():
    if request.method == "POST":
        pinfl = _normalize_pinfl(request.form.get("pinfl"))
        if len(pinfl) != 14:
            flash("PINFL 14 ta raqamdan iborat bo'lishi kerak.", "danger")
            return redirect(url_for("admin.employees"))
        if Employee.query.filter_by(pinfl=pinfl).first():
            flash("Bu PINFL bilan xodim allaqachon mavjud.", "danger")
            return redirect(url_for("admin.employees"))

        emp = Employee(full_name=request.form["full_name"], pinfl=pinfl,
                        position=request.form.get("position"))
        db.session.add(emp)
        log_action("create", "Employee", details=emp.full_name)
        db.session.commit()
        flash(f"{emp.full_name} xodimlar ro'yxatiga qo'shildi.", "success")
        return redirect(url_for("admin.employees"))

    emps = Employee.query.order_by(Employee.full_name).all()
    blocked_count = Customer.query.filter_by(is_blocked=True).count()
    return render_template("admin/employees.html", employees=emps, blocked_count=blocked_count)


@admin_bp.route("/employees/<int:emp_id>/toggle", methods=["POST"])
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def toggle_employee(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    emp.is_active = not emp.is_active
    db.session.commit()
    return redirect(url_for("admin.employees"))


@admin_bp.route("/employees/<int:emp_id>/delete", methods=["POST"])
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def delete_employee(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    log_action("delete", "Employee", entity_id=emp.id, details=emp.full_name)
    db.session.delete(emp)
    db.session.commit()
    flash("Xodim o'chirildi.", "success")
    return redirect(url_for("admin.employees"))


_NAME_HEADER_ALIASES = ("f.i.sh", "fio", "ism familya", "ism", "full_name", "фио", "ф.и.о")
_PINFL_HEADER_ALIASES = ("pinfl", "jshshir", "жшшир", "пинфл", "inn")
_POSITION_HEADER_ALIASES = ("lavozim", "lavozimi", "position", "должность", "должность работника")


def _detect_column(header_row, aliases, fallback_index):
    """Sarlavha qatoridan mos ustunni topadi (nom bo'yicha); topilmasa,
    eski qattiq tartib (0=ism, 1=PINFL, 2=lavozim) bilan orqaga qaytadi."""
    for idx, cell in enumerate(header_row or []):
        if cell and str(cell).strip().lower() in aliases:
            return idx
    return fallback_index


def _coerce_pinfl(value) -> str:
    """Excel katakchasi son (int/float) yoki matn bo'lishidan qat'iy nazar
    PINFL'ni raqamlarga ajratib oladi (masalan floatning ".0" qo'shimchasi
    kabi artefaktlarsiz)."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return _normalize_pinfl(str(value))


@admin_bp.route("/employees/import", methods=["POST"])
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def import_employees():
    """Excel orqali ommaviy import. Ustunlar: F.I.Sh | PINFL | Lavozimi (ixtiyoriy).
    Ustun tartibi sarlavha nomi bo'yicha avtomatik aniqlanadi; topilmasa standart
    tartib (1=F.I.Sh, 2=PINFL, 3=Lavozim) qo'llanadi."""
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Fayl tanlanmadi.", "danger")
        return redirect(url_for("admin.employees"))

    from openpyxl import load_workbook
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        sheet = wb.active
    except Exception:
        flash("Faylni o'qib bo'lmadi. Excel (.xlsx) fayl yuklang.", "danger")
        return redirect(url_for("admin.employees"))

    all_rows = sheet.iter_rows(min_row=1, values_only=True)
    try:
        header_row = next(all_rows)
    except StopIteration:
        flash("Fayl bo'sh.", "danger")
        return redirect(url_for("admin.employees"))

    name_col = _detect_column(header_row, _NAME_HEADER_ALIASES, 0)
    pinfl_col = _detect_column(header_row, _PINFL_HEADER_ALIASES, 1)
    position_col = _detect_column(header_row, _POSITION_HEADER_ALIASES, 2)

    added = 0
    skip_invalid_pinfl = 0
    skip_missing_name = 0
    skip_duplicate = 0
    existing_pinfls = {e.pinfl for e in Employee.query.all()}

    for row in all_rows:
        if not row or all(v is None for v in row):
            continue

        full_name = str(row[name_col]).strip() if name_col < len(row) and row[name_col] else ""
        pinfl = _coerce_pinfl(row[pinfl_col]) if pinfl_col < len(row) else ""
        position = (str(row[position_col]).strip()
                    if position_col < len(row) and row[position_col] else None)

        if not full_name:
            skip_missing_name += 1
            continue
        if len(pinfl) != 14:
            skip_invalid_pinfl += 1
            continue
        if pinfl in existing_pinfls:
            skip_duplicate += 1
            continue

        db.session.add(Employee(full_name=full_name, pinfl=pinfl, position=position))
        existing_pinfls.add(pinfl)
        added += 1

    total_skipped = skip_invalid_pinfl + skip_missing_name + skip_duplicate
    log_action("import", "Employee",
               details=f"added={added}, invalid_pinfl={skip_invalid_pinfl}, "
                       f"missing_name={skip_missing_name}, duplicate={skip_duplicate}")
    db.session.commit()
    flash(
        f"{added} ta xodim qo'shildi. Jami {total_skipped} ta qator o'tkazib yuborildi: "
        f"{skip_invalid_pinfl} ta noto'g'ri/bo'sh PINFL, {skip_missing_name} ta F.I.Sh yo'q, "
        f"{skip_duplicate} ta takroriy PINFL.",
        "success" if added else "danger",
    )
    return redirect(url_for("admin.employees"))


# ---------------------------------------------------------------------------
# Bloklangan murojaatchilar (PINFL tashkilot xodimlari ro'yxatida topilmagan)
# ---------------------------------------------------------------------------
@admin_bp.route("/blocked-customers")
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def blocked_customers():
    customers = Customer.query.filter_by(is_blocked=True).order_by(Customer.created_at.desc()).all()
    return render_template("admin/blocked_customers.html", customers=customers)


@admin_bp.route("/blocked-customers/<int:customer_id>/unblock", methods=["POST"])
@login_required
@roles_required(RoleEnum.SUPER_ADMIN, RoleEnum.ADMINISTRATOR)
def unblock_customer(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    customer.is_blocked = False
    customer.blocked_reason = None
    customer.pinfl_verified = True
    log_action("update", "Customer", entity_id=customer.id, details="unblocked")
    db.session.commit()
    flash(f"{customer.full_name or customer.telegram_id} bloklanishi bekor qilindi.", "success")
    return redirect(url_for("admin.blocked_customers"))


@admin_bp.route("/audit-log")
@login_required
@roles_required(RoleEnum.SUPER_ADMIN)
def audit_log():
    logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(300).all()
    return render_template("admin/audit_log.html", logs=logs)
