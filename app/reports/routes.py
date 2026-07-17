import io
from datetime import datetime

from flask import Blueprint, render_template, send_file, request
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from app.models import ServiceRequest, RequestStatus

reports_bp = Blueprint("reports", __name__)


@reports_bp.route("/")
@login_required
def index():
    return render_template("reports/index.html")


def _filtered_requests():
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    status = request.args.get("status")

    query = ServiceRequest.query
    if date_from:
        query = query.filter(ServiceRequest.created_at >= datetime.strptime(date_from, "%Y-%m-%d"))
    if date_to:
        query = query.filter(ServiceRequest.created_at <= datetime.strptime(date_to, "%Y-%m-%d"))
    if status:
        query = query.filter(ServiceRequest.status == status)
    return query.order_by(ServiceRequest.created_at.desc()).all()


@reports_bp.route("/export/excel")
@login_required
def export_excel():
    reqs = _filtered_requests()

    wb = Workbook()
    ws = wb.active
    ws.title = "Murojaatlar"

    headers = ["№", "Raqam", "Murojatchi", "Telefon", "Bo'linma", "Kategoriya", "Holat", "Ustuvorlik", "Ijrochi",
               "Yaratilgan", "Muddat", "Bajarilgan", "Kechikkanmi"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for i, r in enumerate(reqs, start=1):
        executor = r.current_executor
        ws.append([
            i, r.number, r.customer.full_name or "", r.customer.phone or "",
            r.org_display, r.category.name_uz, r.status.value,
            r.priority.value if r.priority else "",
            executor.full_name if executor else "",
            r.created_at.strftime("%Y-%m-%d %H:%M"),
            r.deadline_at.strftime("%Y-%m-%d %H:%M") if r.deadline_at else "",
            r.completed_at.strftime("%Y-%m-%d %H:%M") if r.completed_at else "",
            "Ha" if r.is_overdue else "Yo'q",
        ])

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"hisobot_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@reports_bp.route("/export/pdf")
@login_required
def export_pdf():
    reqs = _filtered_requests()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [Paragraph("Xizmat ko'rsatish murojaatlari bo'yicha hisobot", styles["Title"]), Spacer(1, 12)]

    data = [["№", "Raqam", "Bo'linma", "Kategoriya", "Holat", "Yaratilgan"]]
    for i, r in enumerate(reqs, start=1):
        data.append([str(i), r.number, r.org_display, r.category.name_uz, r.status.value,
                     r.created_at.strftime("%Y-%m-%d")])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FA")]),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)

    filename = f"hisobot_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buf, as_attachment=True, download_name=filename, mimetype="application/pdf")
